import sqlite3
import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import configparser

# Function to get all tables from a database
def get_all_tables(db_path):
    conn = sqlite3.connect(db_path)
    query = "SELECT name FROM sqlite_master WHERE type='table';"
    tables = pd.read_sql(query, conn)
    conn.close()
    return tables

# Function to check if a table exists in a database
def table_exists(db_path, table_name):
    conn = sqlite3.connect(db_path)
    query = "SELECT name FROM sqlite_master WHERE type='table' AND name=?;"
    result = pd.read_sql(query, conn, params=(table_name,))
    conn.close()
    return not result.empty

# Function to get all data from a table
def get_table_data(db_path, table_name):
    if not table_exists(db_path, table_name):
        raise ValueError(f"Table '{table_name}' does not exist in the database '{db_path}'.")
    
    conn = sqlite3.connect(db_path)
    query = f"SELECT * FROM {table_name}"
    data = pd.read_sql(query, conn)
    conn.close()
    return data

# Load configuration
config = configparser.ConfigParser()
config.read('config.ini')

# Retrieve database paths, table names, and column prefixes from config.ini
db1_path = config['paths']['db1_path']
db2_path = config['paths']['db2_path']
table_name = config['tables']['table_name']
col1_prefix = config['columns']['col1_prefix']
col2_prefix = config['columns']['col2_prefix']
output_path= config['paths']['output_path']

# Retrieve data from the specified table in both databases
data_db1 = get_table_data(db1_path, table_name)
data_db2 = get_table_data(db2_path, table_name)

# Convert data to JSON format
data_json_db1 = data_db1.to_json(orient='records')
data_json_db2 = data_db2.to_json(orient='records')

# Parse JSON data
data1 = json.loads(data_json_db1)
data2 = json.loads(data_json_db2)

# Ensure the tables have the same columns
if set(data_db1.columns) != set(data_db2.columns):
    raise ValueError("Tables do not have the same columns.")

# Initialize comparison results
total_cells = 0
matching_cells = 0
column_matching_counts = {col: 0 for col in data_db1.columns if col != 'ts'}
column_matching_percentages = {}

# Prepare a list to collect the merged rows for the Excel output
merged_rows = []

# Compare each row and column where 'ts' matches
for row1 in data1:
    for row2 in data2:
        if row1['ts'] == row2['ts']:
            merged_row = {'ts': row1['ts']}
            for col in data_db1.columns:
                if col != 'ts':  # Skip 'ts' column for matching percentage calculation
                    merged_row[f'{col}_{col1_prefix}'] = row1[col]
                    merged_row[f'{col}_{col2_prefix}'] = row2[col]
                    total_cells += 1
                    if row1[col] == row2[col]:
                        matching_cells += 1
                        column_matching_counts[col] += 1
                        merged_row[f'{col}_Comparison'] = "Correct"
                    else:
                        merged_row[f'{col}_Comparison'] = "FALSCH"
            merged_rows.append(merged_row)
            break

# Calculate overall matching percentage
overall_matching_percentage = (matching_cells / total_cells) * 100

# Calculate matching percentage for each column
for col in column_matching_counts:
    column_matching_percentages[col] = (column_matching_counts[col] / len(data_db1)) * 100

# Convert merged rows to DataFrame
merged_df = pd.DataFrame(merged_rows)

# Add a row for matching percentages
matching_percentages_row = {'ts': 'Matching Percentage'}
for col in data_db1.columns:
    if col != 'ts':
        matching_percentages_row[f'{col}_{col1_prefix}'] = ""
        matching_percentages_row[f'{col}_{col2_prefix}'] = ""
        matching_percentages_row[f'{col}_Comparison'] = f"{column_matching_percentages[col]:.2f}%"
matching_percentages_df = pd.DataFrame([matching_percentages_row])

# Concatenate the matching percentages row to the merged_data DataFrame
merged_df = pd.concat([merged_df, matching_percentages_df], ignore_index=True)

# Save the result to an Excel file
excel_path = output_path
merged_df.to_excel(excel_path, index=False)

# Load the workbook and select the active sheet
wb = load_workbook(excel_path)
ws = wb.active

# Define the fill for "FALSCH" and "Correct" cells
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# Apply the fills to "FALSCH" and "Correct" cells
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ws.min_column, max_col=ws.max_column):
    for cell in row:
        if cell.value == "FALSCH":
            cell.fill = red_fill
        elif cell.value == "Correct":
            cell.fill = green_fill

# Save the styled workbook
wb.save(excel_path)

# Output results
print(f"Overall Matching Percentage: {overall_matching_percentage:.2f}%")
print(f"Total Number of Compared Cells: {total_cells}")
print("Column Matching Percentages and Counts:")
for col, perc in column_matching_percentages.items():
    print(f"{col}: {perc:.2f}% (Matched: {column_matching_counts[col]} out of {len(data_db1)})")
print(f"Comparison result saved to: {excel_path}")
